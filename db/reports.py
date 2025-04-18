import aiosqlite
import sqlite3
import openpyxl
from openpyxl.styles import Font

DATABASE_PATH = "telegram.db"

async def add_info_in_report(who_open_direction, summary, price, finish_price, company_winner_name, status, who_close_direct):
    who_open_direction = str(who_open_direction)
    async with aiosqlite.connect(DATABASE_PATH) as db:
        await db.execute('''
            INSERT INTO report (who_open_direct, direction_name, start_price, finish_price, company_winner_name, direction_status, who_close_direct)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (who_open_direction, summary, price, finish_price, company_winner_name, status, who_close_direct))
        await db.commit()
        await db.close()
    print("Направление было успешно добавлено в отчет")

async def report_direction_list():
    """
    Генерация отчета по направлениям в формате Excel.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        select_query = '''
        SELECT u.username, d.direction_name, d.direction_description, d.direction_price, d.direction_status, d.winner
        FROM direction d
        JOIN users u ON d.users_id = u.telegram_id
        '''
        cursor = await db.execute(select_query)
        data = await cursor.fetchall()

        # Создание нового Excel-файла
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Запись заголовков с жирным шрифтом
        headers = ['Username', 'Direction Name', 'Direction Description', 'Direction Price', 'Direction Status', 'Winner']
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header).font = openpyxl.styles.Font(bold=True)

        # Запись данных в Excel
        for row_num, row_data in enumerate(data, start=2):  # Начинаем со второй строки, чтобы не перезаписывать заголовки
            for col_num, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Сохранение Excel-файла
        workbook.save("report_direction_list.xlsx")
        print("Отчет по направлениям успешно создан.")

async def report_direction():
    """
    Генерация отчета по направлениям в формате Excel.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        select_query = '''
        SELECT 
            dl.id AS "ID направления",
            u.username AS "Кто открыл направление",
            dl.direction AS "Название направления",
            dl.old_price AS "Начальная цена",
            dl.new_price AS "Окончательная цена",
            dl.company_name AS "Победитель",
            dl.direction_status AS "Статус направления"
        FROM directionList dl
        LEFT JOIN users u ON dl.telegram_id = u.telegram_id
        '''
        cursor = await db.execute(select_query)
        data = await cursor.fetchall()

        # Создание нового Excel-файла
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Запись заголовков с жирным шрифтом
        headers = [
            "ID направления", "Кто открыл направление", "Название направления",
            "Начальная цена", "Окончательная цена", "Победитель",
            "Статус направления"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Запись данных
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                # Преобразуем кортежи в строки, если нужно
                if isinstance(cell_value, tuple):
                    cell_value = ', '.join(map(str, cell_value))
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Сохранение файла
        workbook.save("report_direction.xlsx")
        print("Отчет успешно создан: report_direction.xlsx")

async def update_price(price: str, direction_name_for_close: str, direction_winner_name: str):
    """
    Асинхронное обновление цены в отчете.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
            UPDATE report
            SET finish_price = ?
            WHERE direction_name = ? AND company_winner_name = ?
        '''
        await db.execute(update_query, (price, direction_name_for_close, direction_winner_name))
        await db.commit()
        print(f"Цена для направления '{direction_name_for_close}' обновлена на {price}.")

async def get_last_id_report():
    """
    Асинхронное получение последнего ID отчета.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        select_query = '''
        select id from report
        order by id desc
        limit 1
        '''
        cursor = await db.execute(select_query)
        result = await cursor.fetchone()
        return result[0] if result else None

async def report_close_get_price(direction_name_for_close: str, direction_winner_name: str):
    """
    Асинхронное получение цены для закрытого направления.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        select_query = '''
        SELECT new_price
        FROM directionList
        WHERE direction = ?
        AND company_name = ?
        '''
        cursor = await db.execute(select_query, (direction_name_for_close, direction_winner_name))
        result = await cursor.fetchone()
        return result[0] if result else None
    
async def update_report_status(direction_name: str, status: str):
    """
    Асинхронное обновление статуса направления в отчете.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET direction_status = ?
        WHERE direction_name = ?
        '''
        await db.execute(update_query, (status, direction_name))
        await db.commit()
        await db.close()
        print(f"Статус направления {direction_name} обновлен на {status}.")

async def update_report_winner(direction_name: str, winner_name: str):
    """
    Асинхронное обновление имени победителя в отчете.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET company_winner_name = ?
        WHERE direction_name = ?
        '''
        await db.execute(update_query, (winner_name, direction_name))
        await db.commit()
        await db.close()
        print(f"Победитель для направления {direction_name} обновлен в отчете на {winner_name}.")

async def close_report(direction_name: str):
    """
    Асинхронное закрытие направления в отчете.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET direction_status = 'Close'
        WHERE direction_name = ?
        '''
        await db.execute(update_query, (direction_name,))
        await db.commit()
        await db.close()
        print(f"Направление {direction_name} было закрыто в отчете.")

async def update_company_send_price_to_direction(company_name: str, direction_id: int):
    """
    Асинхронное обновление названия компании, отправившей цену, для направления.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET company_send_price = ?
        WHERE direction_id = ?
        '''
        await db.execute(update_query, (company_name, direction_id))
        await db.commit()
        await db.close()
        print(f"Название компании {company_name} добавлено для направления с ID {direction_id}.")

async def close_direction_set_status(direction_name_for_close: str):
    """
    Асинхронное обновление статуса направления на 'Close' в таблице report.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query_close = '''
        UPDATE report
        SET direction_status = 'Close'
        WHERE direction_name = ?
        '''
        await db.execute(update_query_close, (direction_name_for_close,))
        await db.commit()
        await db.close()
        print(f"Статус направления '{direction_name_for_close}' обновлен на 'Close'.")

async def close_direction_set_who_close_direction(username_who_close: str, direction_name_for_close: str):
    """
    Асинхронное обновление имени пользователя, закрывшего направление, в таблице report.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query_close = '''
        UPDATE report
        SET who_close_direct = ?
        WHERE direction_name = ?
        '''
        await db.execute(update_query_close, (username_who_close, direction_name_for_close))
        await db.commit()
        await db.close()
        print(f"Пользователь {username_who_close} установлен как закрывший направление {direction_name_for_close}.")

async def set_winner_name_report(direction_winner_name: str, direction_name_for_close: str):
    """
    Асинхронное обновление имени победителя для направления в таблице report.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET company_winner_name = ?
        WHERE direction_name = ?
        '''
        await db.execute(update_query, (direction_winner_name, direction_name_for_close))
        await db.commit()
        await db.close()
        print(f"Победитель {direction_winner_name} установлен для направления {direction_name_for_close}.")

async def add_info_in_report_insert_direction_id(direction_id: int, last_id: int):
    """
    Асинхронное обновление ID направления в таблице report.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        update_query = '''
        UPDATE report
        SET direction_id = ?
        WHERE id = ?
        '''
        await db.execute(update_query, (direction_id, last_id))
        await db.commit()
        await db.close()
        print(f"ID направления {direction_id} добавлен в отчет с ID {last_id}.")

async def update_company_send_price_to_direction(company_name, direction_id):
    async with aiosqlite.connect(DATABASE_PATH) as db:
        await db.execute('''
            UPDATE report
            SET company_send_price = ?
            WHERE direction_id = ?
        ''', (company_name, direction_id))
        await db.commit()
        await db.close()
    print("Запись названия компании добавлена")