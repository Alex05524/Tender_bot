import sqlite3
import aiosqlite
import openpyxl
from openpyxl.styles import Font

DATABASE_PATH = "telegram.db"

# Создаем соединение с базой данных
conn = sqlite3.connect('telegram.db')

# Создаем курсор
cursor = conn.cursor()

def insert_direction_list_info(ID, company_name, direction, direction_old_price, direction_new_price, direction_status):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    insert_query = '''
    INSERT INTO directionList (telegram_id, company_name, direction, old_price, new_price, direction_status)
    values (?,?,?,?,?,?)
    '''
    cursor.execute(insert_query, (ID, company_name, direction, direction_old_price, direction_new_price, direction_status))
    conn.commit()
    conn.close()
    print(f"Добавлено новое направление в лист")


def get_direction_without_price(direction):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    select_query = '''
            SELECT direction
            FROM directionList
            WHERE direction = ?
        '''
    cursor.execute(select_query, (direction,))
    directions_without_price = [row[0] for row in cursor.fetchall()]
    conn.close()

    return directions_without_price

def select_my_direction(company_name):
    """
    Возвращает список направлений для указанной компании с активным статусом.
    """
    try:
        conn = sqlite3.connect('telegram.db')
        cursor = conn.cursor()

        select_query = '''
            SELECT direction
            FROM directionList
            WHERE company_name = ?
            AND direction_status = 'active'
        '''
        cursor.execute(select_query, (company_name,))
        directions = [row[0] for row in cursor.fetchall()]
        print(f"Запрос выполнен для компании: {company_name}, найдено направлений: {len(directions)}")
    except sqlite3.Error as e:
        print(f"Ошибка при выполнении запроса select_my_direction: {e}")
        directions = []
    finally:
        conn.close()

    return directions

def select_my_directionlist(direction_name, ID):
    """
    Возвращает информацию о направлении и его новой цене для указанного пользователя.
    """
    try:
        conn = sqlite3.connect('telegram.db')
        cursor = conn.cursor()

        select_query = '''
            SELECT direction, new_price
            FROM directionList
            WHERE direction = ?
            AND telegram_id = ?
            AND direction_status = 'active'
        '''
        cursor.execute(select_query, (direction_name, ID))  # Передаем параметры

        direction = cursor.fetchall()  # Получаем все строки результата
        print(f"Запрос выполнен для направления: {direction_name}, пользователь ID: {ID}, найдено записей: {len(direction)}")
    except sqlite3.Error as e:
        print(f"Ошибка при выполнении запроса select_my_directionlist: {e}")
        direction = []
    finally:
        conn.close()

    return direction

def update_direction_price(direction_name, telegram_id, new_price):
    """
    Обновляет цену для указанного направления и пользователя.
    """
    try:
        conn = sqlite3.connect('telegram.db')
        cursor = conn.cursor()

        update_query = '''
            UPDATE directionList
            SET new_price = ?
            WHERE direction = ?
            AND telegram_id = ?
            AND direction_status = 'active'
        '''
        cursor.execute(update_query, (new_price, direction_name, telegram_id))
        conn.commit()
        print(f"Цена для направления '{direction_name}' успешно обновлена на {new_price}.")
    except sqlite3.Error as e:
        print(f"Ошибка при обновлении цены для направления '{direction_name}': {e}")
    finally:
        conn.close()

def get_my_price_for_direction(company_name,direction_name_list):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    select_query = '''
    select new_price from directionList
    where company_name = ?
    and direction = ?
    and direction_status = 'active'
    '''

    cursor.execute(select_query, (company_name,direction_name_list))
    direction_new_price = [row[0] for row in cursor.fetchall()]
    conn.close()

    return direction_new_price

def get_direction_list_direction(company_name, direction_name):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    select_query = '''
    select * from directionList
    where company_name = ?
    and direction = ?
    and direction_status = 'active'
    '''

    cursor.execute(select_query,(company_name, direction_name))
    direction_list_info = cursor.fetchone()

    return direction_list_info

async def get_my_price_for_direction_by_id(telegram_id: int, direction_name: str):
    """
    Получает текущую цену пользователя (new_price) для указанного направления по telegram_id.
    """
    async with aiosqlite.connect(DATABASE_PATH) as db:
        try:
            query = '''
            SELECT new_price
            FROM directionList
            WHERE telegram_id = ? AND direction = ? AND direction_status = 'active'
            '''
            cursor = await db.execute(query, (telegram_id, direction_name))
            result = await cursor.fetchone()
            return result  # Возвращает new_price
        except Exception as e:
            print(f"Ошибка при выполнении запроса get_my_price_for_direction_by_id: {e}")
            return None

def get_direction_list_direction_by_company(company_name, direction_name):
    """
    Проверяет, отправляла ли компания цену для данного направления.
    """
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    try:
        # Приведение данных к строковому типу
        company_name = str(company_name).strip()
        direction_name = str(direction_name).strip()

        # SQL-запрос для проверки
        select_query = '''
        SELECT * FROM directionList
        WHERE company_name = ?
        AND direction = ?
        AND direction_status = 'active'
        '''

        cursor.execute(select_query, (company_name, direction_name))
        direction_list_info = cursor.fetchone()

        # Логирование результата
        print(f"Список направлений для проверки на отклик:\n{direction_list_info}\nКомпания: {company_name}")

        return direction_list_info
    except sqlite3.Error as e:
        print(f"Ошибка при выполнении запроса get_direction_list_direction_by_company: {e}")
        return None
    finally:
        conn.close()

def close_direction_l(direction_name_for_close):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    update_query1 = '''
        update direction set direction_status = 'Close'
        where direction_name = ?
    '''

    update_query2 = '''
        update directionList set direction_status = 'Close'
        where direction = ?
    '''
    cursor.execute(update_query1, (direction_name_for_close,))
    cursor.execute(update_query2, (direction_name_for_close,))
    conn.commit()
    conn.close()
    print(f"Направление {direction_name_for_close} было закрыто")

def report_close_get_price(direction_name_for_close, direction_winner_name):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    select_query = '''
        select new_price from directionList
        where direction = ? and 
        company_name  = ?
    '''
    cursor.execute(select_query, (direction_name_for_close, direction_winner_name))
    result = cursor.fetchone()
    return result
    
def chek_price_to_report(direction_name_for_close, direction_winner_name):
    conn = sqlite3.connect('telegram.db')
    cursor = conn.cursor()

    select = '''
    select new_price from directionList
    where direction = ?
    and company_name = ?
    and direction_status = 'active'
    '''
    cursor.execute(select, (direction_name_for_close, direction_winner_name))
    result = cursor.fetchall()
    return result
    
async def report_direction_list():
    async with aiosqlite.connect(DATABASE_PATH) as db:
        select_query = '''
        SELECT u.username, d.direction_name, d.direction_description, d.direction_price, d.direction_status, d.winner
        FROM direction d
        JOIN users u ON d.users_id = u.telegram_id
        '''
        async with db.execute(select_query) as cursor:
            data = await cursor.fetchall()

    # Создание нового Excel-файла
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Запись заголовков с жирным шрифтом
    headers = ['Username', 'Direction Name', 'Direction Description', 'Direction Price', 'Direction Status', 'Winner']
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)

    # Запись данных в Excel
    for row_num, row_data in enumerate(data, start=2):  # Начинаем со второй строки, чтобы не перезаписывать заголовки
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Сохранение Excel-файла
    workbook.save("report_direction_list.xlsx")
    print("Отчет успешно создан: report_direction_list.xlsx")